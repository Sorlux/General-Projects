import openpyxl as xl
import calendar
from openpyxl.chart import (ScatterChart, BarChart, Reference, Series)
jan = []
feb = []
mar = []
apr = []
may = []
jun = []
jul = []
aug = []
sep = []
oct = []
nov = []
dec = []
months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug","sep","oct","nov","dec"]

"""
Creates the table that will be used in the excel sheet and the program with chosen categories.

Arguments:
- fullyear (list): a list of all the days in the year in month abbreviated format
- categories (list): a list of categories chosen by the user for budgeting
- sheet (object): an object representing the sheet in an excel workbook
"""

def Table_Creation(fullyear,categories,sheet):
    # Standardizing the excel sheet in this specific way enables the program to function properly.
    cell = 0
    for row in range(2,len(fullyear)+2):
        cell = sheet.cell(row,1)
        cell.value = fullyear[row-2]

    for column in range(2,len(categories)+2):
        cell = sheet.cell(1,column)
        cell.value = categories[column-2]  

    # Sets every value in the table to 0 so that there are no issues due to a cell containg a None type.
    for row in range(2,len(fullyear)+2):

        for column in range(2,len(categories)+2):
            cell = sheet.cell(row,column)
            cell.value = 0

"""
Creates list of all the days in the year in month abbreviated format using a function and the calendar module.
Format was specified so there wouldn't be any ambiguity in what date format the user uses and it is clear and easy to type with.
Used a function and Calendar module instead of hard coding so that in future iterations of the program, it will ask the user for a specific year and account for a leap year.

Returns:
- fullyear (list): a list of all the days in the year in month abbreviated format
"""

def Days_In_The_Year():
    fullyear = []
    obj = calendar.Calendar()
    months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug","sep","oct","nov","dec"]

    for x in range(12):
        temp = []

        for day in obj.itermonthdays(2023,x+1):

            if day != 0:
                temp.append(months[x] + " " + str(day))  

        globals()[months[x]] = temp

        for y in globals()[months[x]]:
            fullyear.append(y)    

    return fullyear

"""
Asks the user for categories to use for budgeting, and creates 4 extra categories for data analysis.

Returns:
- categories (list): a list of categories chosen by the user for budgeting, plus 4 extra categories
"""
def Category_Initialization():
    answer = ""
    categories = []
    i = False
    print("What categories do you wish to use for your budgeting? (if finished adding categories, write 'done')")

    while answer != "y":
        category = input()
        category = category.lower()
        if category in categories:
            print("Sorry!, you already have this category within the list!")
        elif category == "done":
            i = False

            while i != True:
                # Gives the user a chance to view all their categories and change them if needed
                print(f"The categories you have chosen are: {categories} is this correct? (y/n)")
                answer = input()
                answer = answer.lower()

                if answer == "y":
                    print("Thank you! I will create your budget table")
                    i = True
                elif answer == "n":
                    print("Sorry! Could you please give me the names of the categories you wish to use once more? (if finished adding categories, write done)")
                    categories = []
                    i = True
                else:
                    print("Sorry that was an invalid answer please answer with y or n")

        else:
            categories.append(category)

    # Creates 4 extra categories to do data analysis using the information the user gives later on
    categories.append("date total")
    categories.append("month total")
    categories.append("category total")
    categories.append("overall total")

    return categories

"""
Opens an Excel document and switches the sheet to the first sheet in the workbook.

Returns:
- workbook (object): an object representing an Excel workbook
- sheet (object): an object representing the sheet in an Excel workbook
"""
def Set_Workbook_And_Sheet():
    # Want to make sure that the correct excel workbook and sheet are being used so that all the functions work properly.
    workbook = xl.load_workbook("Budget.xlsx")
    current_sheet = workbook.sheetnames
    workbook.active = workbook[current_sheet[0]]
    sheet = workbook[current_sheet[0]]

    return workbook, sheet

"""
Overall function used to house all the other functions related to creating the calendar and the table as a whole.
"""
def Calendar_Creation():
    # Can I remove this?
    fullyear = Days_In_The_Year()
    categories = Category_Initialization()

    workbook = xl.Workbook()
    workbook.save("Budget.xlsx")

    workbook, sheet = Set_Workbook_And_Sheet()

    cell = sheet.cell(1,1)
    cell.value = "Day"

    Table_Creation(fullyear,categories,sheet)

    workbook.save("Budget.xlsx")

"""
Confirms whether the Excel sheet is formatted properly before running the main portion of the program.

Arguments:
- startup (bool): a boolean indicating whether the program is starting up for the first time or not
"""
def Calendar_Confirmation(startup):

    while startup == False:

        # Try caluse was used so that if any error pop us when checking to see if the Excel sheet is formmated properly, it will create a new sheet.
        try:
            workbook, sheet = Set_Workbook_And_Sheet()
            cell = sheet.cell(2,1)

            # Use Jan 1 because no matter what year or categories chosen this cell should always contain this value and is unaffected by any other cell.
            # If this cell is not correct, then the rest of the program will definetely encounter a problem later on.
            if cell.value != "jan 1":
                Calendar_Creation()
            else:
                startup = True
                return startup

        except:
            Calendar_Creation()

"""
Counts the total number of days in the year and returns a list containing all the dates in abbreviated format.
"""
def Count_Days():

    fullyear = []
    workbook, sheet = Set_Workbook_And_Sheet()
    for row in range(2,368):
        cell = sheet.cell(row,1)
        if cell.value != None:
            fullyear.append(cell.value)
    return fullyear


"""
Counts the total number of categories in the table and returns a list containing the name of those categories
"""
def Count_Categories():

    categories = []
    workbook, sheet = Set_Workbook_And_Sheet()
    column = 1
    cell = sheet.cell(1,1)

    while cell.value != None:
        column += 1
        cell = sheet.cell(1,column)
        if cell.value != None:
            categories.append(cell.value)
    return categories

"""
Enables the user to select different options for the specific date they have chosen

Arugments:
cell - The current cell chosen which houses the date and specific category the user is viewing
current_date: The date that is currently being viewed
current_category: The category that is currently being viewed
"""
def Date_Action(cell, current_date, current_category):

    answer = ""

    while answer != "return":
        # Giving the user the immediate information of the cell every loop allows them to see how their actions affect the value directly.
        print(f"The current value you have for {current_date} in category {current_category} is: {cell.value}")
        print("Do you want to (reset) that value or (change) it?")
        answer = input()
        answer.lower()

        if answer == "return":
            continue
        elif answer == "reset":
            cell.value = 0
        elif answer == "change":
            print("Type the amount of money you want to either add or subtract from the current amount")
            money = input()
            money = float(money)
            cell.value = cell.value + money
        else:
            print("Sorry, that answer was invalid, Please try again.")

        Saving()



"""
The user can choose a specific category they want to view for the date they chose

Arguments:
date_index: the index for the current date on the excel sheet
"""
def Date_Selection_Loop(date_index):

    answer = ""
    current_date = fullyear[date_index]
    current_category = []

    while answer != "return":

        print(f"The currently selected date is {current_date}")
        print("Type which category you want to see information for:\n")

        for x in categories:
            print(x)
        print("")

        answer = input()
        answer = answer.lower()

        if answer == "month total" or answer == "category total":
            print("Sorry this option is unavailable when looking at dates!")

        # Before continuing into the rest of the program, will set indices that will continue throughout the program
        # This makes the code used in the rest of the Date option more organized while enabling the use of these indeces with the Calculte_Totals function
        elif answer in categories:
            category_index = categories.index(answer)
            cell = sheet.cell(date_index+2,category_index+2)

            current_category = categories[category_index]
            Date_Action(cell, current_date, current_category)

        # Automatically calculating the totals ensures that the data for totals is instantly updated with the current data
            Calculate_Totals(date_index, "date total")
            Calculate_Totals(date_index, "overall total", True)


"""
Houses the functions related to calculating the totals in the table

Arguments:
interested_index - The index that is causing this change to occur because it was manipulated elsewhere
total_category - The chosen totals category from the 4 categories tracking the total values
overall - Checks to see if the total category is the only category that must be updated
"""
def Calculate_Totals(interested_index, total_category, overall = False,):

    category_index = categories.index(total_category)
    total = 0

# If the date data was not changed, there is no reason to calculate it
    if overall == False:
        for y in range(interested_index+1):
            total = 0
            for x in range(category_index):
                cell = sheet.cell(y+2,x+2)
                total = total + cell.value
            cell = sheet.cell(y+2,category_index+2)
            cell.value = total
    
# Overall needs to always be up to date to be used in data analysis so calculated separately 
    elif overall == True:
        for y in range(fullyear.index("dec 31")+1):
                cell = sheet.cell(y+2,categories.index("date total")+2)
                total = total + cell.value
                cell = sheet.cell(y+2,category_index+2)
                cell.value = total
    # Save all the changes to the workbook
    Saving()

""" 
Overall loop used when the Date option is selected.
Houses all the functions related to the Date category   
""" 
def Date_Loop():
    answer = ""
    while answer != "return":
        print("What date do you want to view? (type help for more information)")
        answer = input()
        answer = answer.lower()

        if answer not in fullyear and answer != "return":
            print("Sorry, that answer was invalid, Plase try again.")
        elif answer == "return":
            continue
        else:
            date_index = fullyear.index(answer)
            Date_Selection_Loop(date_index)

"""
Graphs the data as a Scatter Chart showing how the amount of total money spent has changed throughout the year.
Will add more options for the type of graph and the data being graphed in the future.
"""
def Graphing_Loop():
    yvalues = Reference(sheet,
            min_row= 1, 
            max_row = fullyear.index("dec 31")+1, 
            min_col = categories.index("overall total")+2,
            max_col = categories.index("overall total")+2)

    xvalues = Reference(sheet,
            min_row= 2, 
            max_row = fullyear.index("dec 31")+1,
            min_col = 1,
            max_col = 1)


    chart = ScatterChart()
    chart.height = 10
    chart.width = 20
    chart.title = "Amount of Money Spended Throughout Year"
    chart.style = 5
    chart.x_axis.title = "Days in the Year (Days)"
    chart.y_axis.title = "Money (CAD)"
    series = Series(yvalues, xvalues, title_from_data=True)
    chart.series.append(series)
    sheet.add_chart(chart, "U2")
    Saving()

def Categories_Loop():
    print("Sorry, this function is not available at the moment")

def Month_Loop():
    print("Sorry, this function is not available at the moment")

def Graph_Loop():
    Graphing_Loop()
    Saving()

def Main_Help():
    print("Sorry, this function is not available at the moment")

"""
Main Loop of the program, enabling the user to view and change their data, manipulate it, and create graphs to analyze it
"""
def Main_Loop():
    print("Hello! Welcome to your Expenses of 2023!")
    program = ""

# Enables the program to basically use a switch/case statement where the inputted value determines which function to use.
    switcher = {
        "date" : Date_Loop,
        "categories" : Categories_Loop,
        "month" : Month_Loop,
        "graph" : Graph_Loop,
        "help" : Main_Help,
        "exit" : "exit"
    }

    while program != "exit":
        print(f"Currently you have the Categories: {categories}")
        print("What would you like to do Today? \n")

# Allows the program to go through the keys in the dictionary 
        for x, value in switcher.items():
            print(x)

        print("")
        program = input()
        program = program.lower()
        if program in switcher:
            if program != "exit":
                switcher[program]()
        else:
            print("Sorry! That was an invalid answer, please try again.")
"""
Saves the workbook and checks to see if it is open.
"""
def Saving():
    try:
        workbook.save("Budget.xlsx")
    except PermissionError:
        print("The workbook failed to save because it is currently still open")
        print("Please close the workbook so the program can save any changes you have made to the workbook!")
    except:
        print("An error occured and the workbook was not saved as a result.")



startup = False
# Checks to see if the Excel sheet exists and is formatted correctly. If not, creates a new Excel sheet.
startup = Calendar_Confirmation(startup)

# Set global variables so that all the functions can access their data (Be careful not to change their values while using the program).
fullyear = Count_Days()
categories = Count_Categories()
workbook, sheet = Set_Workbook_And_Sheet()

#If any changes are made on the excel document directly, this will calculate the totals for the entire year before options begin.
Calculate_Totals(fullyear.index("dec 31"),"date total")

Saving()

Main_Loop()
